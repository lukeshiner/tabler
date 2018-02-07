"""This module provides a Table Type for Microsft Excel (.xlsx) files."""

from openpyxl import Workbook, load_workbook

from .basetabletype import BaseTableType


class XLSX(BaseTableType):
    """Table Type for Microsft Excel (.xlsx) files.

    :param str extension: Extension of file to save. Default .xlsx.
    """

    extensions = ['.xlsx']

    def __init__(self, extension='.xlsx'):
        """Consturct :class:`tabler.tabletypes.XLSX`.

        :param str extension: Extension of file to save. Default .xlsx.
        """
        super().__init__(extension)

    def open(self, path):
        """Return header and rows from file.

        :param path: Path to file to be opened.
        :type path: str, pathlib.Path or compatible.
        """
        wb = load_workbook(filename=str(path))
        ws = wb.active
        data = []
        for row in ws:
            data.append([cell.value for cell in row])
        return data[0], data[1:]

    def write(self, table, path):
        """Save data from :class:`tabler.Table` to file.

        :param table: Table to save.
        :type table: :class:`tabler.Table`
        :param path: Path to file to be opened.
        :type path: str, pathlib.Path or compatible.
        """
        wb = Workbook()
        ws = wb.active
        ws.append(table.header)
        for row in table:
            ws.append(row.row)
        wb.save(path)
        print('Writen ' + str(len(table.rows)) + ' lines to file ' + str(path))
